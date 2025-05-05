from django.shortcuts import render
import json
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Ders, DersProgrami, Bolum

# Create your views here.

@csrf_exempt
def ders_ekle(request):
    """Yeni bir ders eklemek için API endpointi."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            yeni_ders = Ders(
                ad=data['ad'],
                kod=data['kod'],
                haftalik_saat=data['haftalik_saat'],
                bolum_id=data['bolum_id'],
                ogretim_uyesi_id=data['ogretim_uyesi_id']
            )
            yeni_ders.save()
            return JsonResponse({"message": "Ders başarıyla eklendi!"}, status=201)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    return JsonResponse({"error": "Geçersiz istek."}, status=405)

def ders_programi_list(request):
    """Ders programlarını listeleyen view."""
    ders_programlari = DersProgrami.objects.all().select_related('ders', 'derslik')
    return render(request, 'program/ders_programi_list.html', {
        'ders_programlari': ders_programlari
    })

def ders_programi_excel(request):
    """Ders programını Excel dosyası olarak dışa aktaran API endpointi."""
    # Excel dosyası oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ders Programı"

    # Stil tanımlamaları
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Renk tanımlamaları
    colors = {
        'pembe': 'FFE6E6',    # Açık pembe
        'mavi': 'E6E6FF',     # Açık mavi
        'yesil': 'E6FFE6',    # Açık yeşil
        'sari': 'FFFFE6',     # Açık sarı
    }

    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 15  # Bölüm
    ws.column_dimensions['B'].width = 15  # Saat
    for col in ['C', 'D', 'E', 'F']:  # Sınıflar için
        ws.column_dimensions[col].width = 35

    # Başlık satırı
    headers = ['Bölüm', 'Saat', '1. Sınıf', '2. Sınıf', '3. Sınıf', '4. Sınıf']
    
    # Saat aralıkları
    saatler = [
        '08:00-09:00', '09:00-10:00', '10:00-11:00', '11:00-12:00',
        '12:00-13:00', '13:00-14:00', '14:00-15:00', '15:00-16:00',
        '16:00-17:00', '17:00-18:00', '18:00-19:00'
    ]

    # Günler
    gunler = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']

    # Bölümleri al
    bolumler = Bolum.objects.all()
    current_row = 1

    for bolum in bolumler:
        # Bölüm başlığı
        bolum_cell = ws.cell(row=current_row, column=1)
        bolum_cell.value = bolum.ad
        bolum_cell.border = border
        bolum_cell.alignment = alignment
        bolum_cell.font = Font(bold=True)
        
        # Her gün için
        for gun in gunler:
            # Gün başlığı
            current_row += 1
            gun_cell = ws.cell(row=current_row, column=1)
            gun_cell.value = gun
            gun_cell.border = border
            gun_cell.alignment = alignment
            gun_cell.font = Font(bold=True)
            
            # Başlık satırı
            current_row += 1
            for col, header in enumerate(headers[1:], 2):  # Bölüm sütununu atlıyoruz
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment
                cell.font = Font(bold=True)

            # Her saat için
            for saat in saatler:
                current_row += 1
                # Saat hücresi
                saat_cell = ws.cell(row=current_row, column=2)
                saat_cell.value = saat
                saat_cell.border = border
                saat_cell.alignment = alignment

                # Her sınıf için (1-4)
                for sinif in range(1, 5):
                    cell = ws.cell(row=current_row, column=sinif + 2)
                    cell.border = border
                    cell.alignment = alignment

                    # Bu bölüm, gün, saat ve sınıfta ders var mı kontrol et
                    dersler = DersProgrami.objects.filter(
                        ders__bolum=bolum,
                        gun=gun,
                        saat=saat,
                        sinif=sinif
                    ).select_related('ders')

                    if dersler.exists():
                        ders = dersler.first()
                        # Sadece dersin adını yazalım
                        cell.value = ders.ders.ad
                        
                        # Sınıflara göre renk belirleme
                        if sinif == 1:
                            cell.fill = PatternFill(start_color=colors['pembe'], end_color=colors['pembe'], fill_type="solid")
                        elif sinif == 2:
                            cell.fill = PatternFill(start_color=colors['mavi'], end_color=colors['mavi'], fill_type="solid")
                        elif sinif == 3:
                            cell.fill = PatternFill(start_color=colors['sari'], end_color=colors['sari'], fill_type="solid")
                        elif sinif == 4:
                            cell.fill = PatternFill(start_color=colors['yesil'], end_color=colors['yesil'], fill_type="solid")

        # Bölümler arası boşluk
        current_row += 2

    # Excel dosyasını kaydet ve gönder
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Ders_Programi.xlsx"'
    wb.save(response)
    return response
